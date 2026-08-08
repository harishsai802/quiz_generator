import os
import io
import csv
import json
import random
import re
from functools import wraps

import openpyxl
from flask import (Flask, render_template, request, redirect, url_for,
                    session, flash, jsonify, Response)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from config import Config
from db import query
from utils.pdf_question_generator import generate_questions_from_pdf

app = Flask(__name__)
app.config.from_object(Config)
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)


# ---------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------
def login_required(role=None):
    """role can be a single role string or a list/tuple of allowed roles."""
    allowed = None
    if role is not None:
        allowed = [role] if isinstance(role, str) else list(role)

    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "user_id" not in session:
                flash("Please log in first.", "error")
                return redirect(url_for("login"))
            if allowed and session.get("role") not in allowed:
                flash("You are not authorized to view that page.", "error")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return wrapper
    return decorator


# Routes reachable by both Admin and approved Teacher accounts
STAFF_ROLES = ["admin", "teacher"]


@app.before_request
def enforce_password_change():
    """Once a bulk-registered student's must_change_password flag is set,
    they're confined to the change-password page (plus logout/static
    assets) until they pick their own password."""
    if session.get("must_change_password") and request.endpoint not in (
        "change_password", "logout", "static"
    ):
        flash("Please set a new password before continuing.", "error")
        return redirect(url_for("change_password"))


# ---------------------------------------------------------------
# Public pages
# ---------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """Self-registration is only available for Teacher accounts (subject
    to admin approval). Student accounts are created in bulk by a
    teacher/admin via the Excel upload (see bulk_register_students) so
    students never see a public "sign up" form."""
    if request.method == "POST":
        name = request.form["name"].strip()
        password = request.form["password"]
        email = request.form.get("email", "").strip().lower()
        if not email:
            flash("Email is required for teacher registration.", "error")
            return redirect(url_for("register"))
        existing = query("SELECT id FROM users WHERE email=%s", (email,), fetchone=True)
        if existing:
            flash("Email already registered.", "error")
            return redirect(url_for("register"))
        pw_hash = generate_password_hash(password)
        query("""INSERT INTO users (name, email, password_hash, role, is_approved)
                 VALUES (%s,%s,%s,'teacher',FALSE)""",
              (name, email, pw_hash), commit=True)
        flash("Registration successful! Your teacher account needs admin approval before you can log in.", "success")
        return redirect(url_for("login"))
    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_id = request.form["login_id"].strip()
        password = request.form["password"]
        user = query("SELECT * FROM users WHERE email=%s OR uucms_id=%s",
                     (login_id.lower(), login_id), fetchone=True)
        if user and check_password_hash(user["password_hash"], password):
            if user["role"] == "teacher" and not user["is_approved"]:
                flash("Your teacher account is still pending admin approval.", "error")
                return redirect(url_for("login"))
            session["user_id"] = user["id"]
            session["name"] = user["name"]
            session["role"] = user["role"]
            session["class_id"] = user.get("class_id")
            session["must_change_password"] = bool(user.get("must_change_password"))
            if session["must_change_password"]:
                flash("This is your first login with an auto-generated password — please set a new one.", "error")
                return redirect(url_for("change_password"))
            if user["role"] in STAFF_ROLES:
                return redirect(url_for("admin_dashboard"))
            return redirect(url_for("student_dashboard"))
        flash("Invalid login ID or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


@app.route("/account/change_password", methods=["GET", "POST"])
@login_required()
def change_password():
    """Any logged-in user (student, teacher, or admin) can change their
    own password here."""
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        user = query("SELECT * FROM users WHERE id=%s", (session["user_id"],), fetchone=True)
        if not user or not check_password_hash(user["password_hash"], current_password):
            flash("Your current password is incorrect.", "error")
            return redirect(url_for("change_password"))

        if len(new_password) < 4:
            flash("New password must be at least 4 characters long.", "error")
            return redirect(url_for("change_password"))

        if new_password != confirm_password:
            flash("New password and confirmation don't match.", "error")
            return redirect(url_for("change_password"))

        query("UPDATE users SET password_hash=%s, must_change_password=FALSE WHERE id=%s",
              (generate_password_hash(new_password), session["user_id"]), commit=True)
        session["must_change_password"] = False
        flash("Password updated successfully.", "success")
        redirect_to = "admin_dashboard" if session.get("role") in STAFF_ROLES else "student_dashboard"
        return redirect(url_for(redirect_to))

    return render_template("change_password.html")


# ---------------------------------------------------------------
# Student dashboard
# ---------------------------------------------------------------
@app.route("/student/dashboard")
@login_required(role="student")
def student_dashboard():
    class_id = session.get("class_id")
    # A quiz with no class_id is open to everyone; otherwise it's only
    # visible to students in that exact class/section.
    if class_id:
        quizzes = query("""
            SELECT * FROM quizzes WHERE class_id IS NULL OR class_id=%s
            ORDER BY created_at DESC
        """, (class_id,), fetch=True)
    else:
        quizzes = query("SELECT * FROM quizzes WHERE class_id IS NULL ORDER BY created_at DESC", fetch=True)
    return render_template("student_dashboard.html", quizzes=quizzes)


# ---------------------------------------------------------------
# Admin dashboard
# ---------------------------------------------------------------
@app.route("/admin/dashboard")
@login_required(role=STAFF_ROLES)
def admin_dashboard():
    quizzes = query("""
        SELECT q.*, c.name AS class_name
        FROM quizzes q LEFT JOIN classes c ON q.class_id = c.id
        ORDER BY q.created_at DESC
    """, fetch=True)

    stats = {
        "quizzes": query("SELECT COUNT(*) AS c FROM quizzes", fetchone=True)["c"],
        "students": query("SELECT COUNT(*) AS c FROM users WHERE role='student'", fetchone=True)["c"],
        "attempts": query("SELECT COUNT(*) AS c FROM attempts", fetchone=True)["c"],
        "violations": query("SELECT COUNT(*) AS c FROM violations", fetchone=True)["c"],
        "classes": query("SELECT COUNT(*) AS c FROM classes", fetchone=True)["c"],
    }

    pending_teachers_count = 0
    if session.get("role") == "admin":
        pending_teachers_count = query(
            "SELECT COUNT(*) AS c FROM users WHERE role='teacher' AND is_approved=FALSE",
            fetchone=True)["c"]

    return render_template("admin_dashboard.html", quizzes=quizzes, stats=stats,
                           pending_teachers_count=pending_teachers_count)


@app.route("/admin/teacher_approvals")
@login_required(role="admin")
def teacher_approvals():
    pending = query(
        "SELECT id, name, email, created_at FROM users WHERE role='teacher' AND is_approved=FALSE ORDER BY created_at",
        fetch=True)
    approved = query(
        "SELECT id, name, email, created_at FROM users WHERE role='teacher' AND is_approved=TRUE ORDER BY created_at DESC",
        fetch=True)
    return render_template("teacher_approvals.html", pending=pending, approved=approved)


@app.route("/admin/teacher_approvals/<int:user_id>/approve", methods=["POST"])
@login_required(role="admin")
def approve_teacher(user_id):
    query("UPDATE users SET is_approved=TRUE WHERE id=%s AND role='teacher'", (user_id,), commit=True)
    flash("Teacher account approved.", "success")
    return redirect(url_for("teacher_approvals"))


@app.route("/admin/teacher_approvals/<int:user_id>/reject", methods=["POST"])
@login_required(role="admin")
def reject_teacher(user_id):
    query("DELETE FROM users WHERE id=%s AND role='teacher' AND is_approved=FALSE", (user_id,), commit=True)
    flash("Teacher registration rejected and removed.", "success")
    return redirect(url_for("teacher_approvals"))


def generate_student_password(uucms_id, name):
    """Auto-generated student password = last 4 digits of the UUCMS ID
    + the first 4 letters of the student's name (lowercase).
    e.g. UUCMS ID U18UZ26S0180 + name "Rahul" -> "0180rahu"."""
    digits = re.sub(r"\D", "", str(uucms_id))
    last4 = digits[-4:] if len(digits) >= 4 else digits.zfill(4)
    letters = re.sub(r"[^A-Za-z]", "", str(name))
    first4 = (letters[:4] if len(letters) >= 4 else letters.ljust(4, "x")).lower()
    return f"{last4}{first4}"


# Header keywords that identify the "UUCMS ID" (or equivalent roll/USN/
# registration number) column in an uploaded Excel sheet. Kept broad because
# different colleges label this column differently.
UUCMS_HEADER_KEYWORDS = (
    "uucms", "usn", "regno", "reg no", "reg.no", "registration",
    "admission", "roll", "enrollment", "enrolment", "student id",
    "studentid", "sats",
)

# A real UUCMS-style ID (e.g. U18UZ26S0180) mixes letters and digits and is
# reasonably long. Plain small integers (1, 2, 3 ... a serial/S.No column)
# must NOT match this, which is what was causing IDs to be stored as "1, 2,
# 3..." instead of the actual UUCMS ID when a sheet had a leading S.No column.
UUCMS_VALUE_PATTERN = re.compile(r"^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9]{6,}$")


def detect_uucms_column(header, data_rows):
    """Figure out which column in the uploaded sheet actually holds the
    UUCMS ID, instead of blindly trusting column A (which is very often a
    serial-number / S.No column full of 1, 2, 3, ...).

    Priority:
      1. A header cell that matches one of UUCMS_HEADER_KEYWORDS.
      2. Otherwise, whichever column's values best match the UUCMS ID shape
         (letters + digits, 6+ chars) across the sample rows.
      3. Falls back to column 0 only if nothing else matches at all.
    Returns (column_index, was_confident: bool).
    """
    by_header = next(
        (i for i, h in enumerate(header) if any(k in h for k in UUCMS_HEADER_KEYWORDS)),
        None,
    )
    if by_header is not None:
        return by_header, True

    sample = [r for r in data_rows[:50] if r]
    if not sample:
        return 0, False

    num_cols = max(len(r) for r in sample)
    best_col, best_score = None, 0.0
    for c in range(num_cols):
        vals = [str(r[c]).strip() for r in sample if c < len(r) and r[c] is not None and str(r[c]).strip() != ""]
        if not vals:
            continue
        matches = sum(1 for v in vals if UUCMS_VALUE_PATTERN.match(v))
        score = matches / len(vals)
        if score > best_score:
            best_score, best_col = score, c

    if best_col is not None and best_score >= 0.5:
        return best_col, True
    return 0, False


def detect_name_column(header, uucms_col):
    by_header = next((i for i, h in enumerate(header) if "name" in h), None)
    if by_header is not None:
        return by_header
    return 1 if uucms_col == 0 else 0


def get_or_create_class(name, teacher_id):
    """Look up a class by name (case-insensitive); create it automatically
    if it doesn't exist yet. Returns the class row."""
    name = name.strip()
    existing = query("SELECT * FROM classes WHERE LOWER(name)=LOWER(%s)", (name,), fetchone=True)
    if existing:
        return existing
    new_id = query("INSERT INTO classes (name, created_by) VALUES (%s,%s)",
                    (name, teacher_id), commit=True)
    return query("SELECT * FROM classes WHERE id=%s", (new_id,), fetchone=True)


@app.route("/admin/bulk_register", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def bulk_register_students():
    """Teacher/Admin uploads an Excel file with UUCMS ID + Name columns,
    and picks (or types a new) class/section that every student in that
    file belongs to. Each row becomes a new student account, linked to
    that class; the password is generated automatically (see
    generate_student_password) and shown once so the teacher can share it
    with the student -- it is never stored in plaintext."""
    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Please upload a valid Excel file (.xlsx).", "error")
            return redirect(url_for("bulk_register_students"))

        # ---- Resolve which class these students belong to ----
        new_class_name = request.form.get("new_class_name", "").strip()
        selected_class_id = request.form.get("class_id", "").strip()
        if new_class_name:
            klass = get_or_create_class(new_class_name, session["user_id"])
        elif selected_class_id:
            klass = query("SELECT * FROM classes WHERE id=%s", (selected_class_id,), fetchone=True)
        else:
            klass = None

        if not klass:
            flash("Please select an existing class or type a new class/section name.", "error")
            return redirect(url_for("bulk_register_students"))

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            flash("Could not read that Excel file. Make sure it's a valid .xlsx file.", "error")
            return redirect(url_for("bulk_register_students"))

        if not rows:
            flash("The Excel file is empty.", "error")
            return redirect(url_for("bulk_register_students"))

        # Detect which columns hold the UUCMS ID and Name. Uses the header
        # row when present; otherwise scans the actual values so a leading
        # S.No/serial column (1, 2, 3...) is never mistaken for the UUCMS ID.
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        has_header = any(k in h for h in header for k in UUCMS_HEADER_KEYWORDS) or any("name" in h for h in header)
        data_rows = rows[1:] if has_header else rows

        uucms_col, confident = detect_uucms_column(header if has_header else [], data_rows)
        name_col = detect_name_column(header if has_header else [], uucms_col)

        if not confident:
            flash("Couldn't confidently detect the UUCMS ID column -- please add a header "
                  "like 'UUCMS ID' to the first row of your Excel file and re-upload to be safe. "
                  "Proceeding with a best guess for now.", "error")

        results = []
        for row in data_rows:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                raw_uucms = row[uucms_col]
                raw_name = row[name_col]
            except IndexError:
                continue
            uucms_id = str(raw_uucms).strip() if raw_uucms is not None else ""
            name = str(raw_name).strip() if raw_name is not None else ""
            if not uucms_id or not name:
                continue

            existing = query("SELECT id FROM users WHERE uucms_id=%s", (uucms_id,), fetchone=True)
            if existing:
                results.append({"uucms_id": uucms_id, "name": name, "password": "-", "status": "Already registered"})
                continue

            password = generate_student_password(uucms_id, name)
            pw_hash = generate_password_hash(password)
            query("""INSERT INTO users (name, uucms_id, password_hash, role, class_id, must_change_password)
                     VALUES (%s,%s,%s,'student',%s,TRUE)""",
                  (name, uucms_id, pw_hash, klass["id"]), commit=True)
            results.append({"uucms_id": uucms_id, "name": name, "password": password, "status": "Registered"})

        if not results:
            flash("No valid student rows found. Make sure the file has a UUCMS ID column and a Name column.", "error")
            return redirect(url_for("bulk_register_students"))

        created = sum(1 for r in results if r["status"] == "Registered")
        flash(f"{created} new student account(s) created in class '{klass['name']}' "
              f"out of {len(results)} row(s) processed.", "success")
        return render_template("bulk_register.html", results=results, classes=classes, klass=klass)

    return render_template("bulk_register.html", results=None, classes=classes, klass=None)


@app.route("/admin/classes", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def manage_classes():
    """List all classes/sections, let staff create one manually, and show
    how many students are in each plus quick performance stats (drives the
    class-wise data views)."""
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Class/section name is required.", "error")
        else:
            get_or_create_class(name, session["user_id"])
            flash(f"Class '{name}' is ready.", "success")
        return redirect(url_for("manage_classes"))

    classes = query("""
        SELECT c.*,
               COUNT(DISTINCT u.id) AS student_count,
               COUNT(a.id) AS attempts_count,
               ROUND(AVG(CASE WHEN a.status='submitted' THEN a.score / a.total_questions * 100 END), 1) AS avg_percent
        FROM classes c
        LEFT JOIN users u ON u.class_id = c.id AND u.role = 'student'
        LEFT JOIN attempts a ON a.student_id = u.id
        GROUP BY c.id
        ORDER BY c.name
    """, fetch=True)
    return render_template("manage_classes.html", classes=classes)


@app.route("/admin/classes/<int:class_id>/students")
@login_required(role=STAFF_ROLES)
def class_roster(class_id):
    klass = query("SELECT * FROM classes WHERE id=%s", (class_id,), fetchone=True)
    if not klass:
        flash("Class not found.", "error")
        return redirect(url_for("manage_classes"))
    students = query("""
        SELECT id, name, uucms_id, created_at FROM users
        WHERE class_id=%s AND role='student' ORDER BY uucms_id
    """, (class_id,), fetch=True)
    other_classes = query("SELECT * FROM classes WHERE id != %s ORDER BY name", (class_id,), fetch=True)
    return render_template("class_roster.html", klass=klass, students=students, other_classes=other_classes)


@app.route("/admin/classes/<int:class_id>/promote", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def promote_class(class_id):
    """Move every student in one class to the next semester/section in a
    single click -- e.g. promote all of '3rd Sem CSE - A' into
    '4th Sem CSE - A'. Their login (UUCMS ID), password, and quiz/score
    history all stay exactly as they are; only their class assignment
    changes, so quizzes scoped to the new class become visible to them
    immediately."""
    klass = query("SELECT * FROM classes WHERE id=%s", (class_id,), fetchone=True)
    if not klass:
        flash("Class not found.", "error")
        return redirect(url_for("manage_classes"))

    if request.method == "POST":
        new_class_name = request.form.get("new_class_name", "").strip()
        target_class_id = request.form.get("target_class_id", "").strip()

        if new_class_name:
            target = get_or_create_class(new_class_name, session["user_id"])
        elif target_class_id:
            target = query("SELECT * FROM classes WHERE id=%s", (target_class_id,), fetchone=True)
        else:
            target = None

        if not target:
            flash("Select an existing class or type a new class/semester name to promote into.", "error")
            return redirect(url_for("promote_class", class_id=class_id))
        if target["id"] == klass["id"]:
            flash("Target class must be different from the current class.", "error")
            return redirect(url_for("promote_class", class_id=class_id))

        moved = query("SELECT COUNT(*) AS c FROM users WHERE class_id=%s AND role='student'",
                       (class_id,), fetchone=True)["c"]
        query("UPDATE users SET class_id=%s WHERE class_id=%s AND role='student'",
              (target["id"], class_id), commit=True)

        flash(f"Promoted {moved} student(s) from '{klass['name']}' to '{target['name']}'.", "success")
        return redirect(url_for("class_roster", class_id=target["id"]))

    other_classes = query("SELECT * FROM classes WHERE id != %s ORDER BY name", (class_id,), fetch=True)
    student_count = query("SELECT COUNT(*) AS c FROM users WHERE class_id=%s AND role='student'",
                           (class_id,), fetchone=True)["c"]
    return render_template("promote_class.html", klass=klass, other_classes=other_classes,
                           student_count=student_count)


@app.route("/admin/students/<int:student_id>/move", methods=["POST"])
@login_required(role=STAFF_ROLES)
def move_student(student_id):
    """Move a single student to a different class -- e.g. a student who
    is repeating a semester and shouldn't be promoted with the rest of
    their class."""
    student = query("SELECT * FROM users WHERE id=%s AND role='student'", (student_id,), fetchone=True)
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("manage_classes"))

    fallback = url_for("class_roster", class_id=student["class_id"]) if student["class_id"] else url_for("manage_classes")

    new_class_name = request.form.get("new_class_name", "").strip()
    target_class_id = request.form.get("target_class_id", "").strip()
    if new_class_name:
        target = get_or_create_class(new_class_name, session["user_id"])
    elif target_class_id:
        target = query("SELECT * FROM classes WHERE id=%s", (target_class_id,), fetchone=True)
    else:
        target = None

    if not target:
        flash("Select or type a class to move this student into.", "error")
        return redirect(fallback)

    query("UPDATE users SET class_id=%s WHERE id=%s", (target["id"], student_id), commit=True)
    flash(f"Moved {student['name']} ({student['uucms_id']}) to '{target['name']}'.", "success")
    return redirect(url_for("class_roster", class_id=target["id"]))


@app.route("/admin/classes/<int:class_id>/export")
@login_required(role=STAFF_ROLES)
def export_class_roster(class_id):
    """Download a class roster (UUCMS ID + Name) as an .xlsx file --
    handy for attendance sheets or handing off to another teacher."""
    klass = query("SELECT * FROM classes WHERE id=%s", (class_id,), fetchone=True)
    if not klass:
        flash("Class not found.", "error")
        return redirect(url_for("manage_classes"))

    students = query("""
        SELECT uucms_id, name, created_at FROM users
        WHERE class_id=%s AND role='student' ORDER BY uucms_id
    """, (class_id,), fetch=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = klass["name"][:31] or "Roster"
    ws.append(["UUCMS ID", "Name", "Registered On"])
    for s in students:
        ws.append([s["uucms_id"], s["name"], str(s["created_at"])])
    for col_cells in ws.columns:
        width = max(len(str(c.value)) for c in col_cells if c.value is not None) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(width, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", klass["name"])
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_roster.xlsx"}
    )


@app.route("/admin/students/<int:student_id>/reset_password", methods=["POST"])
@login_required(role=STAFF_ROLES)
def reset_student_password(student_id):
    """Regenerate a student's password using the same UUCMS-ID/name
    formula and force them to change it on next login. Useful when a
    student forgets their password."""
    student = query("SELECT * FROM users WHERE id=%s AND role='student'", (student_id,), fetchone=True)
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("manage_classes"))

    new_password = generate_student_password(student["uucms_id"], student["name"])
    query("UPDATE users SET password_hash=%s, must_change_password=TRUE WHERE id=%s",
          (generate_password_hash(new_password), student_id), commit=True)
    flash(f"Password for {student['name']} ({student['uucms_id']}) reset to: {new_password} "
          f"— they'll be asked to change it on next login.", "success")
    if student["class_id"]:
        return redirect(url_for("class_roster", class_id=student["class_id"]))
    return redirect(url_for("manage_classes"))


@app.route("/admin/students/<int:student_id>/delete", methods=["POST"])
@login_required(role=STAFF_ROLES)
def delete_student(student_id):
    """Teacher/Admin removes a student account entirely, along with their
    quiz attempts and violation history."""
    student = query("SELECT * FROM users WHERE id=%s AND role='student'", (student_id,), fetchone=True)
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for("manage_classes"))

    class_id = student["class_id"]
    query("DELETE FROM violations WHERE student_id=%s", (student_id,), commit=True)
    query("DELETE FROM attempts WHERE student_id=%s", (student_id,), commit=True)
    query("DELETE FROM users WHERE id=%s", (student_id,), commit=True)

    flash(f"Removed student '{student['name']}' ({student['uucms_id']}).", "success")
    if class_id:
        return redirect(url_for("class_roster", class_id=class_id))
    return redirect(url_for("manage_classes"))


@app.route("/admin/classes/<int:class_id>/delete", methods=["POST"])
@login_required(role="admin")
def delete_class(class_id):
    in_use = query("SELECT COUNT(*) AS c FROM users WHERE class_id=%s", (class_id,), fetchone=True)["c"]
    if in_use:
        flash("Can't delete a class that still has students in it. Move or remove those students first.", "error")
        return redirect(url_for("manage_classes"))
    query("UPDATE quizzes SET class_id=NULL WHERE class_id=%s", (class_id,), commit=True)
    query("DELETE FROM classes WHERE id=%s", (class_id,), commit=True)
    flash("Class deleted.", "success")
    return redirect(url_for("manage_classes"))


@app.route("/admin/upload_pdf", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def upload_pdf():
    if request.method == "POST":
        title = request.form["title"].strip()
        num_questions = int(request.form.get("num_questions", 15))
        questions_per_attempt = int(request.form.get("questions_per_attempt", 10))
        duration_minutes = int(request.form.get("duration_minutes", 15))
        file = request.files.get("pdf_file")

        if not file or not file.filename.lower().endswith(".pdf"):
            flash("Please upload a valid PDF file.", "error")
            return redirect(url_for("upload_pdf"))

        quiz_password = request.form.get("quiz_password", "").strip()
        password_hash = generate_password_hash(quiz_password) if quiz_password else None
        class_id = request.form.get("class_id") or None

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        generated = generate_questions_from_pdf(filepath, num_questions=num_questions)
        if not generated:
            flash("Could not generate questions from this PDF. Try a text-based PDF with more content.", "error")
            return redirect(url_for("upload_pdf"))

        quiz_id = query(
            """INSERT INTO quizzes
               (title, source, created_by, class_id, questions_per_attempt, duration_minutes, access_password)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (title, "auto_pdf", session["user_id"], class_id, questions_per_attempt, duration_minutes, password_hash),
            commit=True)

        for q in generated:
            query("""INSERT INTO questions
                     (quiz_id, question_text, option_a, option_b, option_c, option_d, correct_option)
                     VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                  (quiz_id, q["question_text"], q["option_a"], q["option_b"],
                   q["option_c"], q["option_d"], q["correct_option"]), commit=True)

        flash(f"Quiz '{title}' created with {len(generated)} auto-generated questions!", "success")
        return redirect(url_for("admin_dashboard"))

    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)
    return render_template("upload_pdf.html", classes=classes)


@app.route("/admin/create_quiz", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def create_quiz():
    if request.method == "POST":
        title = request.form["title"].strip()
        questions_per_attempt = int(request.form.get("questions_per_attempt", 10))
        duration_minutes = int(request.form.get("duration_minutes", 15))
        quiz_password = request.form.get("quiz_password", "").strip()
        password_hash = generate_password_hash(quiz_password) if quiz_password else None
        class_id = request.form.get("class_id") or None
        quiz_id = query(
            """INSERT INTO quizzes
               (title, source, created_by, class_id, questions_per_attempt, duration_minutes, access_password)
               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
            (title, "manual", session["user_id"], class_id, questions_per_attempt, duration_minutes, password_hash),
            commit=True)
        return redirect(url_for("add_question", quiz_id=quiz_id))
    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)
    return render_template("create_quiz.html", classes=classes)


# ---------------------------------------------------------------
# Edit / delete an existing quiz (admin, or the teacher who created it)
# ---------------------------------------------------------------
def _quiz_or_404(quiz_id):
    return query("SELECT * FROM quizzes WHERE id=%s", (quiz_id,), fetchone=True)


def _can_manage_quiz(quiz):
    """Admins can manage any quiz; teachers only the ones they created."""
    if not quiz:
        return False
    if session.get("role") == "admin":
        return True
    return quiz["created_by"] == session.get("user_id")


@app.route("/admin/quiz/<int:quiz_id>/edit", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def edit_quiz(quiz_id):
    quiz = _quiz_or_404(quiz_id)
    if not quiz:
        flash("Quiz not found.", "error")
        return redirect(url_for("admin_dashboard"))
    if not _can_manage_quiz(quiz):
        flash("You can only edit quizzes you created.", "error")
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        title = request.form["title"].strip()
        questions_per_attempt = int(request.form.get("questions_per_attempt", 10))
        duration_minutes = int(request.form.get("duration_minutes", 15))
        new_password = request.form.get("quiz_password", "").strip()
        remove_password = request.form.get("remove_password") == "on"
        class_id = request.form.get("class_id") or None

        if remove_password:
            password_hash = None
        elif new_password:
            password_hash = generate_password_hash(new_password)
        else:
            password_hash = quiz["access_password"]  # keep whatever was set before

        query("""UPDATE quizzes
                 SET title=%s, questions_per_attempt=%s, duration_minutes=%s, access_password=%s, class_id=%s
                 WHERE id=%s""",
              (title, questions_per_attempt, duration_minutes, password_hash, class_id, quiz_id), commit=True)
        flash("Quiz updated.", "success")
        return redirect(url_for("admin_dashboard"))

    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)
    return render_template("edit_quiz.html", quiz=quiz, classes=classes)


@app.route("/admin/quiz/<int:quiz_id>/delete", methods=["POST"])
@login_required(role=STAFF_ROLES)
def delete_quiz(quiz_id):
    quiz = _quiz_or_404(quiz_id)
    if not quiz:
        flash("Quiz not found.", "error")
        return redirect(url_for("admin_dashboard"))
    if not _can_manage_quiz(quiz):
        flash("You can only delete quizzes you created.", "error")
        return redirect(url_for("admin_dashboard"))

    # Clean up dependent rows first (works regardless of whether FK cascades are set up)
    query("DELETE FROM violations WHERE quiz_id=%s", (quiz_id,), commit=True)
    query("DELETE FROM attempts WHERE quiz_id=%s", (quiz_id,), commit=True)
    query("DELETE FROM questions WHERE quiz_id=%s", (quiz_id,), commit=True)
    query("DELETE FROM quizzes WHERE id=%s", (quiz_id,), commit=True)

    flash(f"Quiz '{quiz['title']}' and all its data were deleted.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/quiz/<int:quiz_id>/add_question", methods=["GET", "POST"])
@login_required(role=STAFF_ROLES)
def add_question(quiz_id):
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (quiz_id,), fetchone=True)
    if request.method == "POST":
        query("""INSERT INTO questions
                 (quiz_id, question_text, option_a, option_b, option_c, option_d, correct_option)
                 VALUES (%s,%s,%s,%s,%s,%s,%s)""",
              (quiz_id, request.form["question_text"], request.form["option_a"],
               request.form["option_b"], request.form["option_c"], request.form["option_d"],
               request.form["correct_option"].strip().upper()), commit=True)
        flash("Question added.", "success")
        return redirect(url_for("add_question", quiz_id=quiz_id))

    questions = query("SELECT * FROM questions WHERE quiz_id=%s", (quiz_id,), fetch=True)
    return render_template("add_question.html", quiz=quiz, questions=questions)


@app.route("/admin/report")
@login_required(role=STAFF_ROLES)
def admin_report():
    class_filter = request.args.get("class_id", "all")
    class_where = ""
    params = []
    if class_filter and class_filter != "all":
        class_where = " AND u.class_id = %s"
        params.append(class_filter)

    violations = query(f"""
        SELECT v.id, u.name AS student_name, u.email, u.uucms_id, q.title AS quiz_title,
               v.violation_type, v.violation_time
        FROM violations v
        JOIN users u ON v.student_id = u.id
        JOIN quizzes q ON v.quiz_id = q.id
        WHERE 1=1 {class_where}
        ORDER BY v.violation_time DESC
    """, tuple(params), fetch=True)

    summary = query(f"""
        SELECT u.name AS student_name, u.email, COUNT(*) AS violation_count
        FROM violations v JOIN users u ON v.student_id = u.id
        WHERE 1=1 {class_where}
        GROUP BY v.student_id ORDER BY violation_count DESC
    """, tuple(params), fetch=True)

    attempts = query(f"""
        SELECT a.id, u.name AS student_name, q.title AS quiz_title,
               a.score, a.total_questions, a.status, a.started_at, a.submitted_at
        FROM attempts a JOIN users u ON a.student_id = u.id
        JOIN quizzes q ON a.quiz_id = q.id
        WHERE 1=1 {class_where}
        ORDER BY a.started_at DESC
    """, tuple(params), fetch=True)

    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)
    return render_template("admin_report.html", violations=violations, summary=summary,
                           attempts=attempts, classes=classes, class_filter=class_filter)


# ---------------------------------------------------------------
# Quiz taking (personalized + randomized per student)
# ---------------------------------------------------------------
def build_new_attempt(student_id, quiz_id):
    """Pick a random subset & order of questions + shuffled options for this student."""
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (quiz_id,), fetchone=True)
    all_questions = query("SELECT * FROM questions WHERE quiz_id=%s", (quiz_id,), fetch=True)

    n = min(quiz["questions_per_attempt"], len(all_questions))
    chosen = random.sample(all_questions, n)
    random.shuffle(chosen)

    question_ids = [q["id"] for q in chosen]

    # shuffle option order per question, remember mapping so grading is correct
    option_order = {}
    for q in chosen:
        letters = ["A", "B", "C", "D"]
        random.shuffle(letters)
        option_order[str(q["id"])] = letters

    attempt_id = query("""INSERT INTO attempts
        (student_id, quiz_id, question_set, option_order, total_questions, status)
        VALUES (%s,%s,%s,%s,%s,'in_progress')""",
        (student_id, quiz_id, json.dumps(question_ids), json.dumps(option_order), n), commit=True)

    return attempt_id


def get_attempt_payload(attempt_id):
    attempt = query("SELECT * FROM attempts WHERE id=%s", (attempt_id,), fetchone=True)
    question_ids = json.loads(attempt["question_set"])
    option_order = json.loads(attempt["option_order"])

    questions = []
    for qid in question_ids:
        q = query("SELECT * FROM questions WHERE id=%s", (qid,), fetchone=True)
        letters = option_order[str(qid)]
        opts_map = {"A": q["option_a"], "B": q["option_b"], "C": q["option_c"], "D": q["option_d"]}
        display_options = [{"letter": chr(65 + i), "text": opts_map[letters[i]]} for i in range(4)]
        questions.append({
            "id": q["id"],
            "question_text": q["question_text"],
            "options": display_options,
        })
    return attempt, questions


@app.route("/quiz/<int:quiz_id>/start")
@login_required(role="student")
def start_quiz(quiz_id):
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (quiz_id,), fetchone=True)
    if not quiz:
        flash("Quiz not found.", "error")
        return redirect(url_for("student_dashboard"))

    if quiz["class_id"] and quiz["class_id"] != session.get("class_id"):
        flash("This quiz is not available to your class.", "error")
        return redirect(url_for("student_dashboard"))

    if quiz["access_password"] and quiz_id not in session.get("unlocked_quizzes", []):
        return redirect(url_for("quiz_access", quiz_id=quiz_id))

    attempt_id = build_new_attempt(session["user_id"], quiz_id)
    session["current_attempt_id"] = attempt_id
    return redirect(url_for("take_quiz", attempt_id=attempt_id))


@app.route("/quiz/<int:quiz_id>/access", methods=["GET", "POST"])
@login_required(role="student")
def quiz_access(quiz_id):
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (quiz_id,), fetchone=True)
    if not quiz:
        flash("Quiz not found.", "error")
        return redirect(url_for("student_dashboard"))

    if not quiz["access_password"]:
        return redirect(url_for("start_quiz", quiz_id=quiz_id))

    if request.method == "POST":
        entered = request.form.get("quiz_password", "")
        if check_password_hash(quiz["access_password"], entered):
            unlocked = session.get("unlocked_quizzes", [])
            unlocked.append(quiz_id)
            session["unlocked_quizzes"] = unlocked
            return redirect(url_for("start_quiz", quiz_id=quiz_id))
        flash("Incorrect quiz password.", "error")

    return render_template("quiz_access.html", quiz=quiz)


@app.route("/quiz/attempt/<int:attempt_id>")
@login_required(role="student")
def take_quiz(attempt_id):
    attempt, questions = get_attempt_payload(attempt_id)
    if attempt["student_id"] != session["user_id"]:
        flash("Not authorized.", "error")
        return redirect(url_for("student_dashboard"))
    if attempt["status"] == "submitted":
        return redirect(url_for("quiz_result", attempt_id=attempt_id))
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (attempt["quiz_id"],), fetchone=True)
    return render_template("take_quiz.html", attempt=attempt, questions=questions, quiz=quiz)


@app.route("/api/log_violation", methods=["POST"])
@login_required(role="student")
def log_violation():
    data = request.get_json()
    attempt_id = data.get("attempt_id")
    violation_type = data.get("violation_type", "unknown")

    attempt = query("SELECT * FROM attempts WHERE id=%s", (attempt_id,), fetchone=True)
    if not attempt or attempt["student_id"] != session["user_id"]:
        return jsonify({"error": "unauthorized"}), 403

    query("INSERT INTO violations (student_id, quiz_id, attempt_id, violation_type) VALUES (%s,%s,%s,%s)",
          (session["user_id"], attempt["quiz_id"], attempt_id, violation_type), commit=True)

    violation_count = query("SELECT COUNT(*) AS c FROM violations WHERE attempt_id=%s",
                            (attempt_id,), fetchone=True)["c"]

    regenerate = violation_count >= Config.VIOLATION_LIMIT
    new_attempt_id = None
    if regenerate and attempt["status"] == "in_progress":
        query("UPDATE attempts SET status='regenerated' WHERE id=%s", (attempt_id,), commit=True)
        new_attempt_id = build_new_attempt(session["user_id"], attempt["quiz_id"])
        session["current_attempt_id"] = new_attempt_id

    return jsonify({
        "logged": True,
        "violation_count": violation_count,
        "regenerate": regenerate,
        "new_attempt_url": url_for("take_quiz", attempt_id=new_attempt_id) if new_attempt_id else None
    })


@app.route("/quiz/submit", methods=["POST"])
@login_required(role="student")
def submit_quiz():
    attempt_id = int(request.form["attempt_id"])
    attempt = query("SELECT * FROM attempts WHERE id=%s", (attempt_id,), fetchone=True)
    if attempt["student_id"] != session["user_id"] or attempt["status"] != "in_progress":
        flash("This attempt is no longer valid.", "error")
        return redirect(url_for("student_dashboard"))

    question_ids = json.loads(attempt["question_set"])
    score = 0
    for qid in question_ids:
        selected = request.form.get(f"q_{qid}")
        q = query("SELECT correct_option FROM questions WHERE id=%s", (qid,), fetchone=True)
        if selected and selected == q["correct_option"]:
            score += 1

    query("""UPDATE attempts SET score=%s, status='submitted', submitted_at=NOW() WHERE id=%s""",
          (score, attempt_id), commit=True)

    return redirect(url_for("quiz_result", attempt_id=attempt_id))


@app.route("/quiz/result/<int:attempt_id>")
@login_required(role="student")
def quiz_result(attempt_id):
    attempt = query("SELECT * FROM attempts WHERE id=%s", (attempt_id,), fetchone=True)
    if attempt["student_id"] != session["user_id"]:
        flash("Not authorized.", "error")
        return redirect(url_for("student_dashboard"))
    quiz = query("SELECT * FROM quizzes WHERE id=%s", (attempt["quiz_id"],), fetchone=True)
    return render_template("result.html", attempt=attempt, quiz=quiz)


@app.route("/student/my_results")
@login_required(role="student")
def my_results():
    attempts = query("""
        SELECT a.id, q.title AS quiz_title, a.score, a.total_questions,
               a.status, a.started_at, a.submitted_at
        FROM attempts a JOIN quizzes q ON a.quiz_id = q.id
        WHERE a.student_id = %s
        ORDER BY a.started_at DESC
    """, (session["user_id"],), fetch=True)

    violation_count = query(
        "SELECT COUNT(*) AS c FROM violations WHERE student_id=%s",
        (session["user_id"],), fetchone=True)["c"]

    return render_template("my_results.html", attempts=attempts, violation_count=violation_count)


# ---------------------------------------------------------------
# Progress chart + leaderboard + badges
# ---------------------------------------------------------------
@app.route("/student/progress")
@login_required(role="student")
def student_progress():
    attempts = query("""
        SELECT a.id, q.title AS quiz_title, a.score, a.total_questions, a.submitted_at
        FROM attempts a JOIN quizzes q ON a.quiz_id = q.id
        WHERE a.student_id=%s AND a.status='submitted'
        ORDER BY a.submitted_at ASC
    """, (session["user_id"],), fetch=True)

    chart_labels, chart_data = [], []
    for a in attempts:
        pct = round((a["score"] / a["total_questions"]) * 100, 1) if a["total_questions"] else 0
        chart_labels.append(a["quiz_title"])
        chart_data.append(pct)

    total_attempts = len(attempts)
    avg_percent = round(sum(chart_data) / total_attempts, 1) if total_attempts else 0
    best_percent = max(chart_data) if chart_data else 0

    # Improvement = average of the more recent half of attempts minus
    # the average of the earlier half, so students can see their trend.
    improvement = None
    if total_attempts >= 2:
        mid = total_attempts // 2
        first_half = chart_data[:mid] or chart_data[:1]
        second_half = chart_data[mid:] or chart_data[-1:]
        improvement = round((sum(second_half) / len(second_half)) - (sum(first_half) / len(first_half)), 1)

    badges = []
    if total_attempts >= 1:
        badges.append(("🎯", "First Steps", "Completed your first quiz"))
    if total_attempts >= 5:
        badges.append(("📚", "Quiz Master", "Completed 5 or more quizzes"))
    if any(p == 100 for p in chart_data):
        badges.append(("🏆", "Perfectionist", "Scored 100% on a quiz"))
    if total_attempts >= 3 and avg_percent >= 90:
        badges.append(("🌟", "Sharpshooter", "90%+ average score"))
    if improvement is not None and improvement >= 10:
        badges.append(("📈", "Most Improved", "Your recent scores are trending upward"))

    return render_template("student_progress.html", attempts=attempts, chart_labels=chart_labels,
                           chart_data=chart_data, avg_percent=avg_percent, best_percent=best_percent,
                           total_attempts=total_attempts, improvement=improvement, badges=badges)


@app.route("/leaderboard")
@login_required(role=["student", "teacher", "admin"])
def leaderboard():
    # Staff can filter by any class via ?class_id=; students default to
    # their own class but can switch to "All classes" with ?class_id=all.
    class_filter = request.args.get("class_id")
    if class_filter is None and session.get("role") == "student":
        class_filter = str(session.get("class_id")) if session.get("class_id") else "all"

    params = []
    where = "WHERE u.role = 'student'"
    if class_filter and class_filter != "all":
        where += " AND u.class_id = %s"
        params.append(class_filter)

    rows = query(f"""
        SELECT u.id, u.name, u.class_id,
               COUNT(a.id) AS attempts_count,
               ROUND(AVG(a.score / a.total_questions) * 100, 1) AS avg_percent
        FROM users u
        JOIN attempts a ON a.student_id = u.id AND a.status = 'submitted'
        {where}
        GROUP BY u.id, u.name, u.class_id
        ORDER BY avg_percent DESC, attempts_count DESC
        LIMIT 50
    """, tuple(params), fetch=True)

    classes = query("SELECT * FROM classes ORDER BY name", fetch=True)

    rank_icons = {1: "🥇", 2: "🥈", 3: "🥉"}
    leaderboard_rows = []
    my_rank = None
    for i, r in enumerate(rows, start=1):
        badges = []
        if i in rank_icons:
            badges.append((rank_icons[i], "Top " + str(i)))
        if r["avg_percent"] is not None and r["avg_percent"] >= 90:
            badges.append(("🌟", "Sharpshooter"))
        if r["attempts_count"] >= 5:
            badges.append(("📚", "Quiz Master"))
        r = dict(r)
        r["rank"] = i
        r["badges"] = badges
        leaderboard_rows.append(r)
        if r["id"] == session.get("user_id"):
            my_rank = i

    return render_template("leaderboard.html", rows=leaderboard_rows, my_rank=my_rank,
                           classes=classes, class_filter=class_filter or "all")


# ---------------------------------------------------------------
# Admin report - CSV export
# ---------------------------------------------------------------
@app.route("/admin/report/export")
@login_required(role=STAFF_ROLES)
def export_report_csv():
    violations = query("""
        SELECT u.name AS student_name, u.email, q.title AS quiz_title,
               v.violation_type, v.violation_time
        FROM violations v
        JOIN users u ON v.student_id = u.id
        JOIN quizzes q ON v.quiz_id = q.id
        ORDER BY v.violation_time DESC
    """, fetch=True)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student", "Email", "Quiz", "Violation Type", "Time"])
    for v in violations:
        writer.writerow([v["student_name"], v["email"], v["quiz_title"],
                          v["violation_type"], v["violation_time"]])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=violation_report.csv"}
    )


if __name__ == "__main__":
    app.run(debug=True)
